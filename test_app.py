"""
Script para probar la aplicación Streamlit localmente
Descarga el archivo Excel de ejemplo y genera datos de prueba
"""

import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta

def create_sample_excel(filename='MARZO_EJEMPLO.xlsx'):
    """Crea un archivo Excel de ejemplo para pruebas"""

    # Crear datos de ejemplo
    np.random.seed(42)

    montadores = ['AAJ', 'SAE', 'DBJ', 'BSM', 'EMS', 'YBS', 'MRJ']
    clientes = ['DEB', 'Moog Medical', 'Arthro Care', 'Hospira', 'Cardinal']
    productos = [
        '90503564', '55855-001-R1', 'PT00117617', 'CL-M-00131-R1',
        '3012A', '3012AFC-R1', '82026-001', 'PT00106972'
    ]

    num_registros = 100
    start_date = datetime(2026, 3, 1)

    data = []
    for i in range(num_registros):
        fecha = start_date + timedelta(days=np.random.randint(0, 22))
        hora_inicio = f"{np.random.randint(0, 23):02d}:{np.random.randint(0, 60):02d}:00"
        duracion_minutos = np.random.randint(15, 180)
        tiempo_neto = duracion_minutos
        tiempo_programado = np.random.randint(30, 240)

        row = {
            'MES': 'MARZO',
            'DÍA': fecha.strftime('%Y-%m-%d'),
            'MONTADOR': np.random.choice(montadores),
            'NOMBRE DEL MONTADOR': 'Operario Test',
            'PRODUCTO': np.random.choice(productos),
            'CLIENTE': np.random.choice(clientes),
            'INY': np.random.randint(1, 60),
            'MONTAJE (X)': 'X' if np.random.random() > 0.5 else '',
            'DESMONTAJE (X)': 'X' if np.random.random() > 0.5 else '',
            'ID': 13500 + i,
            'HORA DE INICIO': hora_inicio,
            'HORA DE FINALIZACIÓN': f"{np.random.randint(0, 23):02d}:{np.random.randint(0, 60):02d}:00",
            'TIEMPO TOTAL (HORAS)': f"{duracion_minutos/60:.2f}",
            'TIEMPO BRUTO (MIN)': duracion_minutos,
            'TIEMPOS PAROS/MUERTOS': '',
            'TIEMPO NETO': tiempo_neto,
            'TIEMPO PROGRAMADO': tiempo_programado,
        }
        data.append(row)

    df = pd.DataFrame(data)

    # Crear workbook con openpyxl
    wb = Workbook()

    # Agregar datos dummy en filas 1-5
    ws = wb.active
    ws.title = 'DATA'

    # Filas de encabezado (puede ser información de la empresa)
    ws['H2'] = 'Marco Solís / Jefe Soporte de Manufactura'
    ws['H3'] = 'Sergio Ortiz / Gerente de Operaciones'
    ws['H4'] = 'Carolina Oviedo / Gestora de Calidad'

    ws['N2'] = 'Código'
    ws['O2'] = 'F03-P-SM-01'
    ws['N3'] = 'Emisión'
    ws['O3'] = '0'
    ws['N4'] = 'Última Revisión'
    ws['O4'] = datetime.now().strftime('%Y-%m-%d')

    # Agregar encabezados en fila 6
    headers = df.columns.tolist()
    for col, header in enumerate(headers, 1):
        ws.cell(row=6, column=col, value=header)

    # Agregar datos
    for row_idx, row_data in enumerate(df.values, 7):
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    # Ajustar ancho de columnas
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 15

    # Guardar
    wb.save(filename)
    print(f"[OK] Archivo de ejemplo creado: {filename}")
    print(f"[INFO] Registros: {len(df)}")
    print(f"[INFO] Usa este archivo para probar el dashboard")

    return filename

def test_local():
    """Prueba la aplicación localmente"""
    import subprocess
    import sys

    print("[INFO] Iniciando Streamlit localmente...")
    print("[INFO] La app se abrirá en: http://localhost:8501")
    print("[INFO] Carga el archivo MARZO_EJEMPLO.xlsx para ver los datos\n")

    subprocess.run([sys.executable, '-m', 'streamlit', 'run', 'app.py'])

if __name__ == '__main__':
    import sys

    if len(sys.argv) > 1 and sys.argv[1] == 'test':
        # Crear archivo de ejemplo y ejecutar Streamlit
        create_sample_excel()
        test_local()
    else:
        # Solo crear archivo de ejemplo
        create_sample_excel()
        print("\n[INFO] Para ejecutar la aplicacion:")
        print("   streamlit run app.py")
        print("\n[INFO] Para crear datos de ejemplo y ejecutar:")
        print("   python test_app.py test")
